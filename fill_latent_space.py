from pathlib import Path
from os import urandom, getcwd
from datetime import datetime
from copy import deepcopy
from json import dump
from shutil import rmtree

from pandas import read_excel
from numpy import isnan
from sklearn.preprocessing import StandardScaler
from openpyxl import load_workbook
from openpyxl.styles import PatternFill

from machine_learning import Grid, HyperTune
from machine_learning.misc import zip_run_name_files
from machine_learning.featurization import featurize_si_aerogels


def save_predictions(predicted_data):

    raw_data_path = "files/si_aerogels/raw_si_aerogels.xlsx/"
    raw_data_path = str(Path(__file__).parent / raw_data_path)
    wb = load_workbook(filename=raw_data_path)

    for _, row in predicted_data.iterrows():

        # The DataFrame start index at 0, whereas Excel starts at 2 (considering the header also counts as a row)
        # There is an additional header row with specifying info, so add one for that as well
        index = int(row['index']) + 3
        sa = row['Surface Area m2/g']

        # Fill in cell value with calculated surface area
        wb['Comprehensive'][f"GZ{index}"] = sa

        # Color the row to let user know what rows were predicted
        for cell in wb['Comprehensive'][f"A{index}":f"HF{index}"][0]:
            cell.fill = PatternFill("solid", fgColor="fff000")

    output_file = Path(getcwd()) / f"{run_name}_si_aerogels_with_predicted.xlsx"
    wb.save(filename=output_file)


def filter_papers(df):
    df = df.loc[df['Final Gel Type'] == 'Aerogel']

    unique_titles = set(df['Title'].tolist())
    papers_with_unknown_sa = 0
    papers_with_all_unknown_sa = 0
    number_of_aerogels_to_fill = 0
    bad_titles = []  # Titles to remove because all aerogels have unknown surface area

    for unique_title in unique_titles:
        sub_df = df.loc[df['Title'] == unique_title]  # Grab rows with title
        surface_areas = sub_df['Surface Area m2/g'].to_numpy()  # Cast the surface areas to a numpy array
        nan_sa = surface_areas[isnan(surface_areas)]  # Grab all nan values in the surface areas array

        if any(nan_sa):
            papers_with_unknown_sa += 1

        if len(nan_sa) == len(surface_areas):  # If entire numpy array is nan
            papers_with_all_unknown_sa += 1
            bad_titles.append(unique_title)

        elif any(nan_sa):  # If there are nan in array, but the entire array is not nan
            number_of_aerogels_to_fill += len(nan_sa)

    for bad_title in bad_titles:
        df = df.loc[df['Title'] != bad_title]

    general_info = {'Number of unique titles that contain aerogels': len(unique_titles),
                    'Number of papers with unknown surface areas': papers_with_unknown_sa,
                    'Number of papers with all unknown surface areas': papers_with_all_unknown_sa,
                    'Number of aerogels to predict surface area': number_of_aerogels_to_fill,
                    'Titles to be removed with Aerogels because all surface areas are unknown': bad_titles,
                    }

    for key, value in general_info.items():
        print(f"{key}: {value}")

    return df, general_info


def model(training_data, testing_data, validation_percent: float = 0.1):

    # Create training_features
    train_features = training_data.drop(y_column, axis=1)
    train_target = training_data[y_column]

    # Define testing features
    test_features = testing_data

    # Create scalers
    features_scaler = StandardScaler()
    target_scaler = StandardScaler()

    # Scale training/testing features
    train_features = features_scaler.fit_transform(train_features)
    test_features = features_scaler.transform(test_features)

    # Scale train target
    train_target = target_scaler.fit_transform(train_target.to_numpy().reshape(-1, 1)).reshape(-1, )

    # Find HyperParameters
    grid = Grid.make_normal_grid(algorithm)  # Make grid for hyper tuning based on algorithm
    tuner = HyperTune(algorithm, train_features, train_target, grid, opt_iter=10,
                      cv_folds=10)  # Get parameters for hyper tuning
    print("Finding Best HyperParameters...")
    tuner, estimator, params = tuner.hyper_tune(method="random")  # Hyper tuning the model
    tuner.plot_overfit(run_name=run_name)
    tuner.plot_val_pva(run_name=run_name)

    print("Training Model...")
    estimator.fit(train_features, train_target, epochs=20)
    predictions = estimator.predict(test_features)

    # Unscale the predictions
    predictions = target_scaler.inverse_transform(predictions.reshape(-1, 1)).reshape(-1, )
    return predictions


def main(data):
    """
    The parameters that are forced in this function are:
        - Tuned
        - Not grouped
        - Hot-encoded string columns
        - Smart-Values fill NaNs in numerical columns
        - Data not clustered
        - There should only be one y column

    The gaol of this function is to fill the missing gaps in certain properties
        - Examples: Surface area, density, gelation time, etc.

    :return: None
    """
    all_data = deepcopy(data)  # Copy orginial data
    all_data = all_data.reset_index()  # Add column with indexes

    # Remove any papers that have all nan surface area aerogels
    all_data, general_info = filter_papers(all_data)

    # Gather and featurize all data, dropping xerogels in the process
    all_data = featurize_si_aerogels(df=all_data, str_method="one_hot_encode", num_method="smart_values",
                                     y_columns=[y_column], drop_columns=drop_columns, remove_xerogels=True,
                                     drop_rows_missing_y=False, leave_index=True)

    # Grab all rows where a surface area exists
    training_data = all_data.loc[isnan(all_data[y_column]) == False]
    training_data = training_data.drop('index', axis=1)

    # Grab all rows where a surface area does not exists
    data_to_predict = all_data.loc[isnan(all_data[y_column]) == True]
    predicting_indexes = data_to_predict['index']
    data_to_predict = data_to_predict.drop('index', axis=1)
    data_to_predict = data_to_predict.drop(y_column, axis=1)

    # Train the models on the data
    predictions = model(training_data, data_to_predict, validation_percent=0.1)
    data_to_predict[y_column] = predictions
    data_to_predict['index'] = predicting_indexes

    save_predictions(data_to_predict)

    output_file = Path(getcwd()) / f"{run_name}_gen_info.json"
    with open(output_file, 'w') as f:
        dump(general_info, f)


if __name__ == "__main__":
    dataset = r"machine_learning_si_aerogels.csv"
    folder = "si_aerogels"
    file_path = "files/si_aerogels/machine_learning_si_aerogels.xlsx/"

    data_path = str(Path(__file__).parent / file_path)
    data = read_excel(data_path)
    seed = int.from_bytes(urandom(3), "big")  # Generate an actual random number
    algorithm = 'nn'

    now = datetime.now()
    date_string = now.strftime("_%Y%m%d-%H%M%S")
    run_name = f"LatentSpace_{algorithm}_{date_string}_{seed}"

    y_column = 'Surface Area m2/g'
    drop_columns = ['Title', 'Porosity', 'Porosity %', 'Pore Volume cm3/g', 'Average Pore Diameter nm',
                    'Bulk Density g/cm3', 'Young Modulus MPa', 'Crystalline Phase',
                    'Average Pore Size nm', 'Thermal Conductivity W/mK', 'Gelation Time mins']
    main(data)
    zip_run_name_files(run_name=run_name)
